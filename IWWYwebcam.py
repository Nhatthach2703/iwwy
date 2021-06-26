import cv2
import numpy as np
import time

import openpyxl

wb = openpyxl.load_workbook('C:/Users/ALICE/Desktop/sample.xlsx') 

label = ''
count = 1
preLabel = 'open'

cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
net = cv2.dnn.readNet('C:/Users/ALICE/Desktop/YOLO/IWWYsource/yolov3_1800.weights', 'C:/Users/ALICE/Desktop/YOLO/IWWYsource/yolov3.cfg')

classes = []
with open ('C:/Users/ALICE/Desktop/YOLO/IWWYsource/yolo.names', 'r') as f:
    classes = f.read().splitlines()


while True:
    _,img = cap.read()
    height, width, _ = img.shape
    blob = cv2.dnn.blobFromImage(img,1/255, (416,416),(0,0,0), swapRB = True, crop = False)

    net.setInput(blob)
    output_l_n = net.getUnconnectedOutLayersNames()
    lOutputs = net.forward(output_l_n)

    boxes = []
    confidences = []
    class_ids = []

    for output in lOutputs:
        for detection in output:
            scores = detection[5:]
            class_id = np.argmax(scores)
            confidence = scores[class_id]
            if confidence>0.5:

                center_x = int(detection[0]*width)
                center_y = int(detection[1]*height)
                w = int(detection[2]*width)
                h = int(detection[3]*height)

                x = int(center_x - w/2)
                y = int(center_y - h/2)

                boxes.append([x,y,w,h])
                confidences.append((float(confidence)))
                class_ids.append(class_id)

    indexes = cv2.dnn.NMSBoxes(boxes,confidences,0.5,0.4)

    font = cv2.FONT_HERSHEY_PLAIN
    colors = np.random.uniform(0,255,size = (len(boxes),3))

    if len(indexes) > 0:
        for i in indexes.flatten():

            x,y,w,h = boxes[i]
            label = str(classes[class_ids[i]])
            confidence = str(round(confidences[i],2))
            color = colors[i]

            cv2.rectangle(img,(x,y),(x+w, y+h),color,2)
            cv2.putText(img,label+" "+confidence,(x,y-20),font,2,(255,255,255),2)

    cv2.imshow('Dectect', img)

    if label == 'close':
        if preLabel == 'open':
            startT = time.time()
        preLabel = label
    
    if preLabel == 'close':
        endT = time.time()
        if endT-startT>=15:
            nTime =  time.localtime()
            if nTime.tm_hour==7:
                sheet = wb['7h00 - 7h59']
            elif nTime.tm_hour==8:
                sheet = wb['8h00 - 8h59']
            elif nTime.tm_hour==9:
                sheet = wb['9h00 - 9h59']
            elif nTime.tm_hour==10:
                sheet = wb['10h00 - 10h59']
            elif nTime.tm_hour==11:
                sheet = wb['11h00 - 11h59'] 
            elif nTime.tm_hour==12:
                sheet = wb['12h00 - 12h59']
            elif nTime.tm_hour==13:
                sheet = wb['13h00 - 13h59']
            elif nTime.tm_hour==14:
                sheet = wb['14h00 - 14h59']
            elif nTime.tm_hour==15:
                sheet = wb['15h00 - 15h59']
            elif nTime.tm_hour==16:
                sheet = wb['16h00 - 16h59']
            else:
                sheet = wb['else']
            if count == 1:  
                sheet.cell(row=int(sheet.max_row)+1, column=1, value=time.ctime())
                print (time.ctime())
                count = 0

    if label == 'open':
        count = 1
        preLabel = label

    key = cv2.waitKey(1)
    if key ==27:
        wb.save('C:/Users/ALICE/Desktop/sample.xlsx')
        break

cap.release()
cv2.destroyAllWindows()